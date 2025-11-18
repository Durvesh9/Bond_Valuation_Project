import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill

# This file contains the Logic to generate the Excel View

def create_report(bond_results_df, ts_results_df, adf_results, acf_plot_path, output_filename="Bond_Analysis_Report.xlsx"):
    print(f"Generating Excel report: {output_filename}")
    
    try:
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            # --- SHEET 1: BOND VALUATION ---
            bond_results_df.to_excel(writer, sheet_name='Bond Valuation', index=False)
            ws_bonds = writer.sheets['Bond Valuation']
            
            # Format Price as Currency
            price_col = bond_results_df.columns.get_loc('price') + 1
            for row in range(2, len(bond_results_df) + 2):
                ws_bonds.cell(row=row, column=price_col).number_format = '$#,##0.00'

            # Chart 1: Prices
            chart1 = BarChart()
            plot_rows = min(len(bond_results_df), 50)
            chart1.title = f"Bond Prices (First {plot_rows})"
            chart1.y_axis.title = "Price ($)"
            chart1.height, chart1.width = 15, 30
            
            data = Reference(ws_bonds, min_col=price_col, min_row=1, max_row=plot_rows + 1)
            cats = Reference(ws_bonds, min_col=bond_results_df.columns.get_loc('bond_id') + 1, min_row=2, max_row=plot_rows + 1)
            chart1.add_data(data, titles_from_data=True)
            chart1.set_categories(cats)
            chart1.legend = None
            ws_bonds.add_chart(chart1, "J2")

            # --- SHEET 2: TIME SERIES ---
            ts_results_df.to_excel(writer, sheet_name='Time Series Analysis', index=False)
            ws_ts = writer.sheets['Time Series Analysis']
            
            # Format Date
            date_col = ts_results_df.columns.get_loc('date') + 1
            for row in range(2, len(ts_results_df) + 2):
                ws_ts.cell(row=row, column=date_col).number_format = 'yyyy-mm-dd'

            # Chart 2: Rates
            chart2 = LineChart()
            chart2.title = "Interest Rate vs. 20-Day Mean"
            chart2.y_axis.title = "Rate"
            chart2.height, chart2.width = 15, 30
            
            rate_col = ts_results_df.columns.get_loc('rate') + 1
            mean_col = ts_results_df.columns.get_loc('rolling_mean_20d') + 1
            
            c2_data1 = Reference(ws_ts, min_col=rate_col, min_row=1, max_row=len(ts_results_df)+1)
            c2_data2 = Reference(ws_ts, min_col=mean_col, min_row=1, max_row=len(ts_results_df)+1)
            c2_cats = Reference(ws_ts, min_col=date_col, min_row=2, max_row=len(ts_results_df)+1)
            
            chart2.add_data(c2_data1, titles_from_data=True)
            chart2.add_data(c2_data2, titles_from_data=True)
            chart2.set_categories(c2_cats)
            
            chart2.series[0].graphicalProperties.line.solidFill = "0000FF"
            chart2.series[1].graphicalProperties.line.solidFill = "FF0000"
            chart2.series[1].graphicalProperties.line.width = 25000
            chart2.x_axis.tickLblSkip = 60
            ws_ts.add_chart(chart2, "F2")

            # --- SHEET 3: SUMMARY ---
            ws_summary = writer.book.create_sheet('Analysis Summary')
            ws_summary['B2'] = "ADF Test Results"
            ws_summary['B2'].font = Font(bold=True, size=14)
            
            row = 4
            for key, val in adf_results.items():
                ws_summary[f'B{row}'], ws_summary[f'C{row}'] = key, val
                ws_summary[f'B{row}'].font = Font(bold=True)
                if key == 'is_stationary_at_5%':
                    color = "C6EFCE" if val == "Yes" else "FFC7CE"
                    ws_summary[f'C{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                row += 1

            if acf_plot_path:
                ws_summary.add_image(Image(acf_plot_path), 'B14')

        print("Report generated successfully.")
    except Exception as e:
        print(f"Report generation failed: {e}")